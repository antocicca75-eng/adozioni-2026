import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials
from fpdf import FPDF
try:
    from consegne_utils import merge_consegne_lists
except ModuleNotFoundError:
    def _norm_str(v):
        if v is None:
            return ""
        return str(v).strip().upper()

    def _q_int(v):
        try:
            q = int(v)
            return q if q > 0 else 1
        except Exception:
            return 1

    def _consegna_key(item):
        return (
            _norm_str(item.get("t")),
            _norm_str(item.get("e")),
            _norm_str(item.get("sez")),
            _norm_str(item.get("c1")),
            _norm_str(item.get("c2")),
            _norm_str(item.get("c3")),
            _norm_str(item.get("c4")),
            _norm_str(item.get("c5")),
        )

    def merge_consegne_lists(esistenti, nuovi):
        esistenti = esistenti or []
        nuovi = nuovi or []

        mappa = {}
        ordine = []

        def upsert(item):
            if not isinstance(item, dict):
                return
            k = _consegna_key(item)
            if k not in mappa:
                copia = dict(item)
                copia["q"] = _q_int(copia.get("q", 1))
                mappa[k] = copia
                ordine.append(k)
                return

            base = mappa[k]
            base["q"] = _q_int(base.get("q", 1)) + _q_int(item.get("q", 1))

            for campo in ("t", "e", "sez", "c1", "c2", "c3", "c4", "c5"):
                if not base.get(campo) and item.get(campo):
                    base[campo] = item.get(campo)

        for it in esistenti:
            upsert(it)

        for it in nuovi:
            upsert(it)

        return [mappa[k] for k in ordine]


# ==============================================================================
# BLOCCO 1: FUNZIONI CONFIGURAZIONE CONSEGNE (CORRETTO)
# ==============================================================================
def salva_config_consegne(db_dict):
    sh = connetti_google_sheets()
    if sh:
        try:
            try:
                foglio = sh.worksheet("ConfigConsegne")
            except:
                foglio = sh.add_worksheet(title="ConfigConsegne", rows="100", cols="20")

            foglio.clear()
            righe = [["Categoria", "Dati_JSON"]]
            for k, v in db_dict.items():
                # Assicuriamoci che ogni libro salvato mantenga le sue proprietà
                righe.append([k, json.dumps(v)])
            foglio.update(righe)
        except Exception as e:
            st.sidebar.error(f"Errore salvataggio config: {e}")


def carica_config_consegne():
    sh = connetti_google_sheets()
    default = {
        "LETTURE CLASSE PRIMA": [], "LETTURE CLASSE QUARTA": [],
        "SUSSIDIARI DISCIPLINE": [], "INGLESE CLASSE PRIMA": [],
        "INGLESE CLASSE QUARTA": [], "RELIGIONE": []
    }
    db_caricato = {}
    if sh:
        try:
            foglio = sh.worksheet("ConfigConsegne")
            dati = foglio.get_all_records()
            if not dati:
                return default
            for r in dati:
                categoria = r["Categoria"]
                lista_libri = json.loads(r["Dati_JSON"])

                # FIX: Quando carichiamo, verifichiamo che ogni libro abbia la sua 'q'
                # Se manca (vecchi salvataggi), mettiamo 1, altrimenti manteniamo il numero salvato
                for libro in lista_libri:
                    if 'q' not in libro:
                        libro['q'] = 1

                db_caricato[categoria] = lista_libri
        except:
            return default
    return db_caricato if db_caricato else default


# ------------------------------------------------------------------------------


# ==============================================================================
# BLOCCO 2: FUNZIONI STORICO CLOUD
# ==============================================================================
def salva_storico_cloud(storico_dict):
    sh = connetti_google_sheets()
    if sh:
        try:
            try:
                foglio = sh.worksheet("StoricoConsegne")
            except:
                foglio = sh.add_worksheet(title="StoricoConsegne", rows="1000", cols="20")
            foglio.clear()
            righe = [["Plesso", "Dati_JSON"]]
            for plesso, dati in storico_dict.items():
                righe.append([plesso, json.dumps(dati)])
            foglio.update(righe)
        except Exception as e:
            st.sidebar.error(f"Errore salvataggio storico: {e}")


def carica_storico_cloud():
    sh = connetti_google_sheets()
    storico_caricato = {}
    if sh:
        try:
            foglio = sh.worksheet("StoricoConsegne")
            dati = foglio.get_all_records()
            for r in dati:
                storico_caricato[r["Plesso"]] = json.loads(r["Dati_JSON"])
        except:
            pass
    return storico_caricato


# ------------------------------------------------------------------------------

def salva_ritiri_cloud(ritiri_dict):
    sh = connetti_google_sheets()
    if sh:
        try:
            try:
                foglio = sh.worksheet("StoricoRitiri")
            except:
                foglio = sh.add_worksheet(title="StoricoRitiri", rows="1000", cols="20")
            foglio.clear()
            righe = [["Plesso", "Dati_JSON"]]
            for plesso, dati in ritiri_dict.items():
                righe.append([plesso, json.dumps(dati)])
            foglio.update(righe)
        except Exception as e:
            st.sidebar.error(f"Errore salvataggio ritiri: {e}")


def carica_ritiri_cloud():
    sh = connetti_google_sheets()
    ritiri_caricati = {}
    if sh:
        try:
            foglio = sh.worksheet("StoricoRitiri")
            dati = foglio.get_all_records()
            for r in dati:
                ritiri_caricati[r["Plesso"]] = json.loads(r["Dati_JSON"])
        except:
            pass
    return ritiri_caricati

def aggiungi_ritiri(plesso, tipo, items):
    if "storico_ritiri" not in st.session_state:
        st.session_state.storico_ritiri = {}
    if plesso not in st.session_state.storico_ritiri:
        st.session_state.storico_ritiri[plesso] = {}
    if tipo not in st.session_state.storico_ritiri[plesso]:
        st.session_state.storico_ritiri[plesso][tipo] = []
    destinazione = st.session_state.storico_ritiri[plesso][tipo]
    for nuovo in items:
        q_new = int(nuovo.get('q', 0))
        unito = False
        for es in destinazione:
            if es.get('t') == nuovo.get('t') and es.get('e') == nuovo.get('e'):
                es['q'] = int(es.get('q', 0)) + q_new
                unito = True
                break
        if not unito:
            destinazione.append(nuovo.copy())


# ==============================================================================
# BLOCCO 3: COSTANTI E SETTAGGI PAGINA
# ==============================================================================
DB_FILE = "dati_adozioni.csv"
CONFIG_FILE = "anagrafiche.xlsx"
ID_FOGLIO = "1Ah5_pucc4b0ziNZxqo0NRpHwyUvFrUEggIugMXzlaKk"

st.set_page_config(page_title="Adozioni 2026", layout="wide", page_icon="📚")


# ------------------------------------------------------------------------------


from fpdf import FPDF
import io

class PDF_CONSEGNA(FPDF):
    def __init__(self, logo_data=None):
        super().__init__(orientation='L', unit='mm', format='A4')
        self.logo_path = "logo.jpg"

    def rounded_rect(self, x, y, w, h, r, style='', corners='1234'):
        k = self.k
        hp = self.h
        if style == 'F':
            op = 'f'
        elif style == 'FD' or style == 'DF':
            op = 'B'
        else:
            op = 'S'
        my_arc = 4 / 3 * (pow(2, 0.5) - 1)
        self._out(f'{(x + r) * k:.2f} {(hp - y) * k:.2f} m')
        xc = x + w - r
        yc = y + r
        self._out(f'{xc * k:.2f} {(hp - y) * k:.2f} l')
        if '2' in corners:
            self._arc(xc + r * my_arc, yc - r, xc + r, yc - r * my_arc, xc + r, yc)
        else:
            self._out(f'{(x + w) * k:.2f} {(hp - y) * k:.2f} l')
        xc = x + w - r
        yc = y + h - r
        self._out(f'{(x + w) * k:.2f} {(hp - yc) * k:.2f} l')
        if '3' in corners:
            self._arc(xc + r, yc + r * my_arc, xc + r * my_arc, yc + r, xc, yc + r)
        else:
            self._out(f'{(x + w) * k:.2f} {(hp - (y + h)) * k:.2f} l')
        xc = x + r
        yc = y + h - r
        self._out(f'{xc * k:.2f} {(hp - (y + h)) * k:.2f} l')
        if '4' in corners:
            self._arc(xc - r * my_arc, yc + r, xc - r, yc + r * my_arc, xc - r, yc)
        else:
            self._out(f'{x * k:.2f} {(hp - (y + h)) * k:.2f} l')
        xc = x + r
        yc = y + r
        self._out(f'{x * k:.2f} {(hp - yc) * k:.2f} l')
        if '1' in corners:
            self._arc(xc - r, yc - r * my_arc, xc - r * my_arc, yc - r, xc, yc - r)
        else:
            self._out(f'{x * k:.2f} {(hp - y) * k:.2f} l')
        self._out(op)

    def _arc(self, x1, y1, x2, y2, x3, y3):
        h = self.h
        self._out(
            f'{x1 * self.k:.2f} {(h - y1) * self.k:.2f} {x2 * self.k:.2f} {(h - y2) * self.k:.2f} {x3 * self.k:.2f} {(h - y3) * self.k:.2f} c'
        )

    def disegna_modulo(self, x_offset, libri, categoria, p, ins, sez, data_m):
        img_w = 70
        img_x = x_offset + (148.5 - img_w) / 2
        img_y = 8
        box_h = 32
        box_w = img_w + 6
        try:
            self.image(self.logo_path, x=img_x, y=img_y + 2, w=img_w)
            self.set_line_width(0.3)
            self.rounded_rect(img_x - 3, img_y, box_w, box_h, 3)
        except:
            self.rounded_rect(img_x - 3, img_y, box_w, box_h, 3)
            self.set_font('Arial', 'I', 7)
            self.text(img_x + 10, img_y + 15, "Logo non trovato")

        self.set_y(46)
        self.set_x(x_offset + 10)
        self.set_fill_color(235, 235, 235)
        self.rounded_rect(x_offset + 10, 46, 128, 8, 2, 'DF')
        self.set_font('Arial', 'B', 10)
        self.cell(128, 8, f"{str(categoria).upper()}", border=0, ln=1, align='C')

        self.set_x(x_offset + 10)
        self.set_fill_color(245, 245, 245)
        self.set_font('Arial', 'B', 8)
        self.cell(75, 7, 'TITOLO DEL TESTO', border=1, align='C', fill=True)
        self.cell(23, 7, 'CLASSE', border=1, align='C', fill=True)
        self.cell(30, 7, 'EDITORE', border=1, ln=1, align='C', fill=True)

        self.set_font('Arial', '', 8)
        for i, lib in enumerate(libri[:15]):
            self.set_x(x_offset + 10)
            self.cell(75, 7, f" {str(lib['t'])[:40]}", border=1, align='L')
            is_quaderni_vacanze = "QUADERNI VACANZE" in str(categoria).upper()
            if is_quaderni_vacanze:
                raw = [
                    str(lib.get('c1', '')).strip().upper(),
                    str(lib.get('c2', '')).strip().upper(),
                    str(lib.get('c3', '')).strip().upper(),
                    str(lib.get('c4', '')).strip().upper(),
                    str(lib.get('c5', '')).strip().upper(),
                ]
                classi = [c for c in raw if c]
                while len(classi) < 5:
                    classi.append("")
                for w, val in [(4.6, classi[0]), (4.6, classi[1]), (4.6, classi[2]), (4.6, classi[3]), (4.6, classi[4])]:
                    self.cell(w, 7, val, border=1, align='C')
            else:
                raw = [
                    str(lib.get('c1', '')).strip().upper(),
                    str(lib.get('c2', '')).strip().upper(),
                    str(lib.get('c3', '')).strip().upper(),
                ]
                classi = [c for c in raw if c]
                while len(classi) < 3:
                    classi.append("")
                self.cell(7.6, 7, classi[0], border=1, align='C')
                self.cell(7.6, 7, classi[1], border=1, align='C')
                self.cell(7.8, 7, classi[2], border=1, align='C')
            self.cell(30, 7, str(lib.get('e', ''))[:18], border=1, ln=1, align='C')

        self.set_y(155)
        self.set_x(x_offset + 10)
        self.set_fill_color(240, 240, 240)
        self.rounded_rect(x_offset + 10, 155, 128, 7, 1.5, 'DF')
        self.set_font('Arial', 'B', 9)
        self.cell(128, 7, ' DETTAGLI DI CONSEGNA', border=0, ln=1)

        campi = [("PLESSO:", p), ("INSEGNANTE:", ins), ("CLASSE:", sez), ("DATA:", data_m)]
        for label, val in campi:
            self.set_x(x_offset + 10)
            self.set_font('Arial', 'B', 8)
            self.cell(35, 6.5, label, border=1, align='L')
            self.set_font('Arial', '', 8)
            t_v = str(val).upper() if val and val != "- SELEZIONA PLESSO -" else ""
            self.cell(93, 6.5, t_v, border=1, ln=1, align='L')

def genera_pdf_due_copie(libri, categoria, plesso, insegnante, classe, data_modulo):
    pdf = PDF_CONSEGNA()
    pdf.set_auto_page_break(False)
    pdf.add_page()
    pdf.disegna_modulo(0, libri, categoria, plesso, insegnante, classe, data_modulo)
    pdf.disegna_modulo(148.5, libri, categoria, plesso, insegnante, classe, data_modulo)
    return io.BytesIO(pdf.output(dest='S').encode('latin1'))

# ==============================================================================
# BLOCCO 5: CONNESSIONE GOOGLE DRIVE E BACKUP
# ==============================================================================
def connetti_google_sheets():
    try:
        scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        json_info = json.loads(st.secrets["gspread"]["json_data"], strict=False)
        if "private_key" in json_info:
            json_info["private_key"] = json_info["private_key"].replace("\\n", "\n")
        creds = Credentials.from_service_account_info(json_info, scopes=scope)
        client_gs = gspread.authorize(creds)
        sh = client_gs.open_by_key(ID_FOGLIO)
        return sh
    except Exception as e:
        st.error(f"⚠️ Errore connessione Cloud: {e}")
        return None


def backup_su_google_sheets(df_da_salvare):
    sh = connetti_google_sheets()
    if sh:
        try:
            foglio = sh.worksheet("Adozioni_DB")
            foglio.clear()
            dati = [df_da_salvare.columns.values.tolist()] + df_da_salvare.fillna("").values.tolist()
            foglio.update(dati)
            return True
        except Exception as e:
            st.sidebar.error(f"Errore scrittura Cloud: {e}")
            return False
    return False


# ------------------------------------------------------------------------------


# ==============================================================================
# BLOCCO 6: STILI CSS, CACHE E CATALOGO LIBRI
# ==============================================================================
st.markdown("""
    <style>
    [data-testid="stDataEditor"] thead tr th { background-color: #004a99 !important; color: white !important; }
    .stApp { background-color: #ffffff; }
    .totale-box { padding: 20px; background-color: #e8f0fe; border-radius: 10px; border: 1px solid #004a99; margin-top: 15px; }
    </style>
    """, unsafe_allow_html=True)


@st.cache_data(ttl=3600)
def get_catalogo_libri():
    sh = connetti_google_sheets()
    if sh:
        try:
            df = pd.DataFrame(sh.worksheet("Catalogo").get_all_records())
            return df.fillna("")
        except:
            pass
    if os.path.exists(CONFIG_FILE):
        try:
            df = pd.read_excel(CONFIG_FILE, sheet_name="ListaLibri")
            df.columns = [c.strip() for c in df.columns]
            return df.fillna("")
        except:
            return pd.DataFrame()
    return pd.DataFrame()


@st.cache_data(ttl=3600)
def get_lista_plessi():
    sh = connetti_google_sheets()
    if sh:
        try:
            df = pd.DataFrame(sh.worksheet("Plesso").get_all_records())
            return sorted(df.iloc[:, 0].dropna().unique().tolist())
        except:
            pass
    if os.path.exists(CONFIG_FILE):
        try:
            df = pd.read_excel(CONFIG_FILE, sheet_name="Plesso")
            return sorted(df.iloc[:, 0].dropna().unique().tolist())
        except:
            return []
    return []


def aggiungi_libro_a_excel(t, m, e, a):
    try:
        # Salva su Google Sheets (foglio "Catalogo")
        sh = connetti_google_sheets()
        if sh:
            foglio_catalogo = sh.worksheet("Catalogo")
            foglio_catalogo.append_row([t, m, e, a])
            st.cache_data.clear()
            return True
        else:
            st.error("❌ Impossibile connettersi a Google Sheets")
            return False
    except Exception as e:
        st.error(f"❌ Errore salvataggio: {e}")
        return False


def aggiorna_libro_catalogo(row_idx, t, m, e, a):
    try:
        sh = connetti_google_sheets()
        if sh:
            foglio_catalogo = sh.worksheet("Catalogo")
            foglio_catalogo.update(f"A{row_idx}:D{row_idx}", [[t, m, e, a]])
            st.cache_data.clear()
            return True
        else:
            st.error("❌ Impossibile connettersi a Google Sheets")
            return False
    except Exception as ex:
        st.error(f"❌ Errore aggiornamento: {ex}")
        return False


def elimina_libro_catalogo(row_idx):
    try:
        sh = connetti_google_sheets()
        if sh:
            r = int(row_idx)
            if r <= 1:
                st.error("❌ Impossibile eliminare l'intestazione.")
                return False
            foglio_catalogo = sh.worksheet("Catalogo")
            foglio_catalogo.delete_rows(r)
            st.cache_data.clear()
            return True
        else:
            st.error("❌ Impossibile connettersi a Google Sheets")
            return False
    except Exception as ex:
        st.error(f"❌ Errore eliminazione: {ex}")
        return False

# ------------------------------------------------------------------------------


# ==============================================================================
# BLOCCO 7: STATO SESSIONE E INIZIALIZZAZIONE
# ==============================================================================
catalogo = get_catalogo_libri()
if not catalogo.empty:
    elenco_titoli = sorted([str(x) for x in catalogo.iloc[:, 0].unique() if str(x).strip() != ""])
    elenco_materie = sorted([str(x) for x in catalogo.iloc[:, 1].unique() if str(x).strip() != ""])
    elenco_editori = sorted([str(x) for x in catalogo.iloc[:, 2].unique() if str(x).strip() != ""])
    elenco_agenzie = sorted([str(x) for x in catalogo.iloc[:, 3].unique() if str(x).strip() != ""])
else:
    elenco_titoli = elenco_materie = elenco_editori = elenco_agenzie = []

elenco_plessi = get_lista_plessi()

if "pagina" not in st.session_state:
    st.session_state.pagina = "Inserimento"

if 'db_consegne' not in st.session_state:
    st.session_state.db_consegne = carica_config_consegne()
if 'lista_consegne_attuale' not in st.session_state:
    st.session_state.lista_consegne_attuale = []


def reset_ricerca():
    st.session_state.r_attiva = False
    st.session_state.ft = []
    st.session_state.fa = []
    st.session_state.fp = []
    st.session_state.fm = []
    st.session_state.fe = []
    st.session_state.fsag = "TUTTI"

def ordina_tipologie(tipologie):
    ordine = [
        "LETTURE CLASSE PRIMA",
        "LETTURE CLASSE QUARTA",
        "SUSSIDIARI DISCIPLINE",
        "INGLESE CLASSE PRIMA",
        "INGLESE CLASSE QUARTA",
        "RELIGIONE",
        "QUADERNI VACANZE CLASSE PRIMA",
        "QUADERNI VACANZE CLASSE SECONDA",
        "QUADERNI VACANZE CLASSE TERZA",
        "QUADERNI VACANZE CLASSE QUARTA",
        "QUADERNI VACANZE CLASSE QUINTA",
        "QUADERNI VACANZE INGLESE",
    ]

    def _n(v):
        return str(v).strip().upper()

    mappa = {}
    for t in (tipologie or []):
        mappa[_n(t)] = t

    scelti = []
    usati = set()
    for o in ordine:
        if o in mappa:
            scelti.append(mappa[o])
            usati.add(o)

    rimanenti = [mappa[k] for k in sorted([k for k in mappa.keys() if k not in usati])]
    return scelti + rimanenti


# ------------------------------------------------------------------------------


# ==============================================================================
# BLOCCO 8: SIDEBAR (NAVIGAZIONE MENU)
# ==============================================================================
with st.sidebar:
    st.title("🧭 MENU")
    if st.button("➕ NUOVA ADOZIONE", use_container_width=True):
        st.session_state.pagina = "Inserimento";
        st.rerun()

    if st.button("✏️ MODIFICA ADOZIONE", use_container_width=True):
        st.session_state.pagina = "Modifica";
        st.rerun()

    if st.button("🆕 AGGIUNGI A CATALOGO", use_container_width=True):
        st.session_state.pagina = "NuovoLibro";
        st.rerun()

    if st.button("✏️ MODIFICA LIBRO", use_container_width=True):
        st.session_state.pagina = "ModificaLibro";
        st.rerun()

    if st.button("📊 REGISTRO COMPLETO", use_container_width=True):
        st.session_state.pagina = "Registro";
        st.rerun()

    if st.button("🔍 PIVOT ADOZIONI", use_container_width=True):
        st.session_state.pagina = "Ricerca";
        st.rerun()

    if st.button("📄 MODULO CONSEGNE", use_container_width=True):
        st.session_state.pagina = "Consegne";
        st.rerun()

    if st.button("📚 COLLANE CONSEGNATE", use_container_width=True):
        st.session_state.pagina = "Storico";
        st.rerun()
    if st.button("📦 COLLANE RITIRATE", use_container_width=True):
        st.session_state.pagina = "Ritirate";
        st.rerun()

    if st.button("🔍 RICERCA COLLANE", use_container_width=True):
        st.session_state.pagina = "Ricerca Collane"
        st.rerun()

    if st.button("📊 TABELLONE STATO", use_container_width=True):
        st.session_state.pagina = "Tabellone Stato";
        st.rerun()
    st.markdown("---")
# ------------------------------------------------------------------------------


# ==============================================================================
# BLOCCO 9: PAGINA CONSEGNE (STAMPA DOPPIA E PDF)
# ==============================================================================
if st.session_state.pagina == "Consegne":
    st.header("📄 Generazione Moduli Consegna")

    if "storico_consegne" not in st.session_state:
        st.session_state.storico_consegne = carica_storico_cloud()
    if "storico_ritiri" not in st.session_state:
        st.session_state.storico_ritiri = carica_ritiri_cloud()

    def reset_consegne_totale():
        st.session_state.lista_consegne_attuale = []
        st.session_state.last_cat = None
        st.rerun()


    ctr = st.session_state.get('reset_ctr', 0)
    actr = st.session_state.get('add_ctr', 0)

    col_p, col_c = st.columns(2)
    plessi_scelti = col_p.multiselect("Seleziona Plesso/i:", elenco_plessi, key=f"p_sel_{ctr}")
    p_scelto = plessi_scelti[0] if len(plessi_scelti) == 1 else ""

    tipologie_fisse_menu = [
        "INGLESE CLASSE PRIMA",
        "INGLESE CLASSE QUARTA",
        "QUADERNI VACANZE CLASSE PRIMA",
        "QUADERNI VACANZE CLASSE SECONDA",
        "QUADERNI VACANZE CLASSE TERZA",
        "QUADERNI VACANZE CLASSE QUARTA",
        "QUADERNI VACANZE CLASSE QUINTA",
        "QUADERNI VACANZE INGLESE"
    ]

    basi = [
        "- SELEZIONA -",
        "TUTTE LE TIPOLOGIE",
        "SELEZIONE MULTIPLA",
    ] + [t for t in tipologie_fisse_menu if t in st.session_state.db_consegne]

    altre = [k for k in st.session_state.db_consegne.keys() if k not in tipologie_fisse_menu]
    cat_scelta = col_c.selectbox("Tipologia Libri:", basi + sorted(altre), key=f"c_sel_{ctr}")
    tipologie_scelte = []

    with st.container(border=True):
        c_nt, c_btn = st.columns([0.75, 0.25])
        nuova_tip = c_nt.text_input("Nuova tipologia di libri", key=f"nuova_tip_{ctr}")
        if c_btn.button("➕ CREA", use_container_width=True, key=f"btn_crea_tip_{ctr}"):
            nome = str(nuova_tip).strip().upper()
            vietate = {"- SELEZIONA -", "TUTTE LE TIPOLOGIE", "SELEZIONE MULTIPLA"}
            if not nome:
                st.warning("Inserisci un nome per la nuova tipologia.")
            elif nome in vietate:
                st.warning("Nome tipologia non valido.")
            elif nome in st.session_state.db_consegne:
                st.warning("Questa tipologia esiste già.")
            else:
                st.session_state.db_consegne[nome] = []
                salva_config_consegne(st.session_state.db_consegne)
                st.session_state.lista_consegne_attuale = []
                st.session_state.last_cat = nome
                nuovo_ctr = st.session_state.get("reset_ctr", 0) + 1
                st.session_state.reset_ctr = nuovo_ctr
                st.session_state[f"c_sel_{nuovo_ctr}"] = nome
                st.rerun()

    if cat_scelta in st.session_state.db_consegne:
        del_tipo_key = f"del_tipo_conf_{ctr}"
        del_nome_key = f"del_tipo_nome_{ctr}"
        b_del = st.button("🗑️ ELIMINA TIPOLOGIA", use_container_width=True, key=f"btn_del_tipo_{ctr}")
        if b_del:
            st.session_state[del_tipo_key] = True
            st.session_state[del_nome_key] = cat_scelta
            st.rerun()

        if st.session_state.get(del_tipo_key) and st.session_state.get(del_nome_key) == cat_scelta:
            st.warning(f"Confermi eliminazione della tipologia: {cat_scelta}?")
            c1, c2 = st.columns(2)
            if c1.button("✅ CONFERMA ELIMINA", use_container_width=True, key=f"btn_del_tipo_ok_{ctr}"):
                if cat_scelta in st.session_state.db_consegne:
                    del st.session_state.db_consegne[cat_scelta]
                    salva_config_consegne(st.session_state.db_consegne)
                st.session_state.pop(del_tipo_key, None)
                st.session_state.pop(del_nome_key, None)
                st.session_state.lista_consegne_attuale = []
                st.session_state.last_cat = None
                st.session_state.reset_ctr = st.session_state.get('reset_ctr', 0) + 1
                st.rerun()
            if c2.button("Annulla", use_container_width=True, key=f"btn_del_tipo_no_{ctr}"):
                st.session_state.pop(del_tipo_key, None)
                st.session_state.pop(del_nome_key, None)
                st.rerun()

    if cat_scelta == "TUTTE LE TIPOLOGIE":
        st.info("💡 Assegnazione massiva selezionata.")
        st.session_state.lista_consegne_attuale = []
        st.session_state.last_cat = "TUTTE"

    elif cat_scelta == "SELEZIONE MULTIPLA":
        st.info("💡 Seleziona più tipologie da assegnare al plesso.")
        st.session_state.lista_consegne_attuale = []
        st.session_state.last_cat = "MULTI"
        tipologie_scelte = st.multiselect(
            "Seleziona una o più tipologie:",
            sorted(list(st.session_state.db_consegne.keys())),
            key=f"multi_tip_{ctr}",
        )

    elif cat_scelta != "- SELEZIONA -" and st.session_state.get('last_cat') != cat_scelta:
        caricati = list(st.session_state.db_consegne.get(cat_scelta, []))
        for voce in caricati: voce['q'] = 1
        st.session_state.lista_consegne_attuale = caricati
        st.session_state.last_cat = cat_scelta

    if cat_scelta not in ["- SELEZIONA -", "TUTTE LE TIPOLOGIE", "SELEZIONE MULTIPLA"]:
        st.markdown("---")
        for i, lib in enumerate(st.session_state.lista_consegne_attuale):
            if 'q' not in lib: lib['q'] = 1
            c_info, c_qta, c_del = st.columns([0.6, 0.3, 0.1])
            is_quaderni_vacanze = "QUADERNI VACANZE" in str(cat_scelta).upper()
            if is_quaderni_vacanze:
                raw = [
                    str(lib.get("c1", "")).strip(),
                    str(lib.get("c2", "")).strip(),
                    str(lib.get("c3", "")).strip(),
                    str(lib.get("c4", "")).strip(),
                    str(lib.get("c5", "")).strip(),
                ]
                classi = [c for c in raw if c]
                while len(classi) < 5:
                    classi.append("")
                c_info.info(f"{lib['t']} | {lib['e']} | Classi: {classi[0]} {classi[1]} {classi[2]} {classi[3]} {classi[4]}")
            else:
                raw = [str(lib.get("c1", "")).strip(), str(lib.get("c2", "")).strip(), str(lib.get("c3", "")).strip()]
                classi = [c for c in raw if c]
                while len(classi) < 3:
                    classi.append("")
                c_info.info(f"{lib['t']} | {lib['e']} | Classi: {classi[0]} {classi[1]} {classi[2]}")
            m1, v1, p1 = c_qta.columns([1, 1, 1])
            if m1.button("➖", key=f"m_{cat_scelta}_{i}"):
                if lib['q'] > 1: lib['q'] -= 1; st.rerun()
            v1.markdown(f"<p style='text-align:center; font-weight:bold; font-size:18px;'>{lib['q']}</p>",
                        unsafe_allow_html=True)
            if p1.button("➕", key=f"p_{cat_scelta}_{i}"):
                lib['q'] += 1;
                st.rerun()
            if c_del.button("❌", key=f"del_{cat_scelta}_{i}"):
                st.session_state.lista_consegne_attuale.pop(i);
                st.rerun()

        col_btns = st.columns(2)
        if col_btns[0].button("💾 REGISTRA LISTA BASE", use_container_width=True):
            lista_da_salvare = []
            for item in st.session_state.lista_consegne_attuale:
                nuovo_item = item.copy();
                nuovo_item['q'] = 1;
                lista_da_salvare.append(nuovo_item)
            st.session_state.db_consegne[cat_scelta] = lista_da_salvare
            salva_config_consegne(st.session_state.db_consegne)
            st.success("Configurazione salvata!")

        if col_btns[1].button("🗑️ SVUOTA SCHERMATA", use_container_width=True):
            st.session_state.reset_ctr = st.session_state.get('reset_ctr', 0) + 1
            reset_consegne_totale()

        with st.expander("➕ Cerca e Aggiungi Libro"):
            df_cat = get_catalogo_libri()
            if not df_cat.empty:
                scelta_libro = st.selectbox("Seleziona libro:", ["- CERCA TITOLO -"] + sorted(
                    df_cat.iloc[:, 0].astype(str).unique().tolist()), key=f"sk_{actr}")
                if scelta_libro != "- CERCA TITOLO -":
                    dati_libro = df_cat[df_cat.iloc[:, 0] == scelta_libro].iloc[0]
                    is_quaderni_vacanze = "QUADERNI VACANZE" in str(cat_scelta).upper()
                    if is_quaderni_vacanze:
                        c_sez, c1, c2, c3, c4, c5, _ = st.columns([1.2, 1, 1, 1, 1, 1, 2])
                        sez_in = c_sez.text_input("Sezione", key=f"sez_{actr}")
                        c1in = c1.text_input("Classe", max_chars=2, key=f"in1_{actr}")
                        c2in = c2.text_input("Classe ", max_chars=2, key=f"in2_{actr}")
                        c3in = c3.text_input("Classe  ", max_chars=2, key=f"in3_{actr}")
                        c4in = c4.text_input("Classe   ", max_chars=2, key=f"in4_{actr}")
                        c5in = c5.text_input("Classe    ", max_chars=2, key=f"in5_{actr}")
                    else:
                        c_sez, c1, c2, c3, _ = st.columns([1.2, 1, 1, 1, 4])
                        sez_in = c_sez.text_input("Sezione", key=f"sez_{actr}")
                        c1in = c1.text_input("Classe", max_chars=2, key=f"in1_{actr}")
                        c2in = c2.text_input("Classe ", max_chars=2, key=f"in2_{actr}")
                        c3in = c3.text_input("Classe  ", max_chars=2, key=f"in3_{actr}")
                        c4in = ""
                        c5in = ""
                    if st.button("Conferma Aggiunta", key=f"btn_add_{actr}", use_container_width=True):
                        raw = [str(c1in).strip(), str(c2in).strip(), str(c3in).strip(), str(c4in).strip(), str(c5in).strip()]
                        classi = [c for c in raw if c]
                        while len(classi) < (5 if is_quaderni_vacanze else 3):
                            classi.append("")
                        if not is_quaderni_vacanze:
                            classi = classi[:3]
                        st.session_state.lista_consegne_attuale.append({
                            "t": str(dati_libro.iloc[0]).upper(), "e": str(dati_libro.iloc[2]).upper(),
                            "q": 1,
                            "c1": classi[0], "c2": classi[1], "c3": classi[2],
                            "c4": classi[3] if is_quaderni_vacanze else "",
                            "c5": classi[4] if is_quaderni_vacanze else "",
                            "sez": sez_in
                        })
                        st.session_state.add_ctr = st.session_state.get('add_ctr', 0) + 1;
                        st.rerun()

    st.markdown("---")
    d1, d2 = st.columns(2)
    docente = d1.text_input("Insegnante ricevente", key=f"doc_{ctr}")
    data_con = d2.text_input("Data di consegna", key=f"dat_{ctr}")
    classe_man = d1.text_input("Classe specifica", key=f"cla_{ctr}")

    col_print, col_conf = st.columns(2)

    # --- PARTE FINALE BLOCCO 9: PDF E CONFERMA ---
    col_print, col_conf = st.columns(2)

    if cat_scelta not in ["TUTTE LE TIPOLOGIE", "- SELEZIONA -", "SELEZIONE MULTIPLA"]:
        if col_print.button("🖨️ GENERA PDF", use_container_width=True):
            if len(plessi_scelti) > 1:
                st.warning("Per generare il PDF seleziona un solo plesso (oppure nessuno).")
            elif st.session_state.lista_consegne_attuale:
                plesso_pdf = plessi_scelti[0] if len(plessi_scelti) == 1 else ""
                # Carichiamo il file fisico logo.jpg per evitare l'errore NameError
                logo_per_pdf = None
                if os.path.exists("logo.jpg"):
                    with open("logo.jpg", "rb") as f:
                        logo_per_pdf = f.read()

                # Creazione PDF con il logo corretto
                pdf = PDF_CONSEGNA(logo_data=logo_per_pdf)
                pdf.add_page()
                pdf.disegna_modulo(0, st.session_state.lista_consegne_attuale, cat_scelta, plesso_pdf, docente,
                                   classe_man, data_con)
                pdf.dashed_line(148.5, 0, 148.5, 210, 0.5)
                pdf.disegna_modulo(148.5, st.session_state.lista_consegne_attuale, cat_scelta, plesso_pdf, docente,
                                   classe_man, data_con)
                st.download_button("📥 SCARICA PDF", bytes(pdf.output()), "consegna.pdf", "application/pdf")

    # --- TASTO CONFERMA SISTEMATO ---
    if col_conf.button("✅ CONFERMA CONSEGNA", use_container_width=True):
        if not plessi_scelti:
            st.warning("Seleziona almeno un plesso.")
        else:
            if cat_scelta == "SELEZIONE MULTIPLA" and not tipologie_scelte:
                st.warning("Seleziona almeno una tipologia.")
            else:
                for plesso in plessi_scelti:
                    if plesso not in st.session_state.storico_consegne:
                        st.session_state.storico_consegne[plesso] = {}

                    if cat_scelta == "SELEZIONE MULTIPLA":
                        for k in tipologie_scelte:
                            v = st.session_state.db_consegne.get(k, [])
                            lista_clean = []
                            for item in v:
                                nuovo = item.copy()
                                nuovo['q'] = 1
                                lista_clean.append(nuovo)
                            esistenti = st.session_state.storico_consegne[plesso].get(k, [])
                            st.session_state.storico_consegne[plesso][k] = merge_consegne_lists(esistenti, lista_clean)

                    elif cat_scelta == "TUTTE LE TIPOLOGIE":
                        for k, v in st.session_state.db_consegne.items():
                            lista_clean = []
                            for item in v:
                                nuovo = item.copy()
                                nuovo['q'] = 1
                                lista_clean.append(nuovo)
                            esistenti = st.session_state.storico_consegne[plesso].get(k, [])
                            st.session_state.storico_consegne[plesso][k] = merge_consegne_lists(esistenti, lista_clean)

                    else:
                        lista_con_quantita_esatte = [item.copy() for item in st.session_state.lista_consegne_attuale]
                        esistenti = st.session_state.storico_consegne[plesso].get(cat_scelta, [])
                        st.session_state.storico_consegne[plesso][cat_scelta] = merge_consegne_lists(
                            esistenti, lista_con_quantita_esatte
                        )

                if cat_scelta == "SELEZIONE MULTIPLA":
                    st.success("Assegnazione multipla completata!")
                elif cat_scelta == "TUTTE LE TIPOLOGIE":
                    st.success("REGISTRAZIONE MASSIVA COMPLETATA!")
                else:
                    st.success("Consegna registrata con successo!")

                salva_storico_cloud(st.session_state.storico_consegne)
                reset_consegne_totale()

# ==============================================================================
# BLOCCO 10: PAGINA STORICO (REGISTRO CARICO PLESSI) - MODIFICA TASTO AGGIORNA
# ==============================================================================
elif st.session_state.pagina == "Storico":
    st.subheader("📚 Registro Libri in Carico ai Plessi")

    if "storico_ritiri" not in st.session_state: st.session_state.storico_ritiri = carica_ritiri_cloud()

    if not st.session_state.get("storico_consegne"):
        st.info("Nessuna consegna registrata.")
    else:
        elenco_plessi_storico = sorted(list(st.session_state.storico_consegne.keys()))

        tipologie_obbligatorie = [
            "LETTURE CLASSE PRIMA",
            "LETTURE CLASSE QUARTA",
            "SUSSIDIARI DISCIPLINE",
            "INGLESE CLASSE PRIMA",
            "INGLESE CLASSE QUARTA",
            "RELIGIONE",
            "QUADERNI VACANZE CLASSE PRIMA",
            "QUADERNI VACANZE CLASSE SECONDA",
            "QUADERNI VACANZE CLASSE TERZA",
            "QUADERNI VACANZE CLASSE QUARTA",
            "QUADERNI VACANZE CLASSE QUINTA",
        ]

        def _norm_tip(v):
            return str(v).strip().upper()

        with st.expander("🧾 Verifica tipologie obbligatorie", expanded=False):
            plessi_check = st.multiselect(
                "🏫 Plessi da controllare",
                elenco_plessi_storico,
                default=elenco_plessi_storico,
                key="chk_plessi_obbl",
            )
            mostra_completi = st.checkbox("Mostra anche plessi completi", value=False, key="chk_show_ok")

            righe = []
            for plesso in plessi_check:
                presenti = {_norm_tip(k) for k in (st.session_state.storico_consegne.get(plesso, {}) or {}).keys()}
                mancanti = [t for t in tipologie_obbligatorie if _norm_tip(t) not in presenti]
                if mostra_completi or mancanti:
                    righe.append({
                        "Plesso": plesso,
                        "N Mancanti": len(mancanti),
                        "Tipologie Mancanti": ", ".join(mancanti),
                    })

            df_chk = pd.DataFrame(righe)
            if df_chk.empty:
                st.success("Tutte le scuole selezionate risultano complete.")
            else:
                df_chk = df_chk.sort_values(by=["N Mancanti", "Plesso"], ascending=[False, True])
                st.dataframe(df_chk, use_container_width=True, hide_index=True)
                try:
                    buf = io.BytesIO()
                    engine = None
                    try:
                        import openpyxl
                        engine = "openpyxl"
                    except Exception:
                        try:
                            import xlsxwriter
                            engine = "xlsxwriter"
                        except Exception:
                            engine = None
                    if engine is None:
                        raise RuntimeError("Nessun engine Excel disponibile")
                    with pd.ExcelWriter(buf, engine=engine) as writer:
                        df_chk.to_excel(writer, index=False, sheet_name="Controllo")
                    st.download_button(
                        "📤 ESPORTA EXCEL CONTROLLO",
                        data=buf.getvalue(),
                        file_name=f"controllo_tipologie_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as ex:
                    st.warning(f"Export Excel non disponibile: {ex}")

        scuole_selezionate = st.multiselect("🔍 Seleziona Plesso/i:", elenco_plessi_storico, key="sel_plessi_storico")
        
        if not scuole_selezionate:
            st.info("Seleziona uno o più plessi per visualizzare le collane consegnate.")
        else:
            for plesso in scuole_selezionate:
                with st.expander(f"🏫 PLESSO: {plesso.upper()}", expanded=True):
                    if st.button(f"📦 RITIRA INTERO PLESSO: {plesso}", key=f"bulk_plesso_{plesso}",
                                 use_container_width=True):
                        for tipo, items in st.session_state.storico_consegne[plesso].items():
                            aggiungi_ritiri(plesso, tipo, items)
                        del st.session_state.storico_consegne[plesso]
                        salva_storico_cloud(st.session_state.storico_consegne);
                        salva_ritiri_cloud(st.session_state.storico_ritiri);
                        st.rerun()
                    if st.button(f"🧨 RESET PLESSO: {plesso}", key=f"reset_plesso_{plesso}", use_container_width=True):
                        if plesso in st.session_state.storico_consegne: del st.session_state.storico_consegne[plesso]
                        if plesso in st.session_state.storico_ritiri: del st.session_state.storico_ritiri[plesso]
                        salva_storico_cloud(st.session_state.storico_consegne);
                        salva_ritiri_cloud(st.session_state.storico_ritiri);
                        st.rerun()

                    per_tipo = st.session_state.storico_consegne[plesso]
                    open_tipo_key = f"open_tipo_consegnate_{plesso}"
                    if open_tipo_key not in st.session_state:
                        st.session_state[open_tipo_key] = None

                    # Forziamo l'allineamento a sinistra del contenuto del bottone
                    st.markdown("""
                        <style>
                        div[data-testid="stButton"] > button {
                            display: flex !important;
                            justify-content: flex-start !important;
                            text-align: left !important;
                            padding-left: 15px !important;
                        }
                        div[data-testid="stButton"] > button > div {
                            width: 100% !important;
                            display: flex !important;
                            justify-content: flex-start !important;
                        }
                        div[data-testid="stButton"] > button > div > p {
                            width: 100% !important;
                            text-align: left !important;
                            margin: 0 !important;
                        }
                        </style>
                        """, unsafe_allow_html=True)

                    for tipo in ordina_tipologie(per_tipo.keys()):
                        is_open = st.session_state.get(open_tipo_key) == tipo
                        freccia = "🔽" if is_open else "▶️"

                        if st.button(f"{freccia} 📘 {tipo.upper()}", key=f"open_tipo_{plesso}_{tipo}", use_container_width=True):
                            st.session_state[open_tipo_key] = None if is_open else tipo
                            st.rerun()

                        if is_open:
                            with st.container(border=True):
                                c_tip1, c_tip2 = st.columns(2)
                                if c_tip1.button("🧨 RESET TIPOLOGIA", key=f"reset_tipo_{plesso}_{tipo}", use_container_width=True):
                                    if plesso in st.session_state.storico_consegne and tipo in st.session_state.storico_consegne[plesso]:
                                        del st.session_state.storico_consegne[plesso][tipo]
                                        if not st.session_state.storico_consegne[plesso]:
                                            del st.session_state.storico_consegne[plesso]
                                    salva_storico_cloud(st.session_state.storico_consegne)
                                    st.rerun()
                                if c_tip2.button("📦 RITIRA TUTTO", key=f"bulk_tipo_{plesso}_{tipo}", use_container_width=True):
                                    aggiungi_ritiri(plesso, tipo, per_tipo[tipo])
                                    del st.session_state.storico_consegne[plesso][tipo]
                                    if not st.session_state.storico_consegne[plesso]:
                                        del st.session_state.storico_consegne[plesso]
                                    salva_storico_cloud(st.session_state.storico_consegne)
                                    salva_ritiri_cloud(st.session_state.storico_ritiri)
                                    st.rerun()

                                lista_libri = list(per_tipo[tipo])
                                for i, lib in enumerate(lista_libri):
                                    qta_salvata = int(lib.get('q', 1))
                                    col_titolo, col_qta, col_adott, col_ritiro, col_del = st.columns([0.35, 0.10, 0.15, 0.30, 0.10])
                                    col_titolo.markdown(f"**{lib['t']}**<br><small>{lib['e']}</small>", unsafe_allow_html=True)
                                    col_qta.write(f"Q.tà: {qta_salvata}")

                                    with col_adott:
                                        if st.button("🌟 ADOTTATO", key=f"adott_{plesso}_{tipo}_{i}"):
                                            st.session_state.adozione_da_storico = {
                                                "plesso": plesso,
                                                "titolo": lib.get("t", ""),
                                                "editore": lib.get("e", ""),
                                                "tipologia": tipo,
                                            }
                                            st.session_state.pagina = "Inserimento"
                                            st.rerun()

                                    with col_ritiro:
                                        q_rit = st.number_input("Ritira", min_value=1, max_value=max(1, qta_salvata),
                                                                value=max(1, qta_salvata), key=f"qrit_{plesso}_{tipo}_{i}",
                                                                label_visibility="collapsed")

                                        if st.button("🔄 AGGIORNA CARICO", key=f"btn_rit_{plesso}_{tipo}_{i}"):
                                            rit_item = lib.copy()
                                            rit_item['q'] = q_rit
                                            aggiungi_ritiri(plesso, tipo, [rit_item])

                                            lib['q'] = qta_salvata - q_rit
                                            if lib['q'] <= 0: per_tipo[tipo].pop(i)

                                            if not st.session_state.storico_consegne[plesso][tipo]: del \
                                                st.session_state.storico_consegne[plesso][tipo]
                                            if not st.session_state.storico_consegne[plesso]: del \
                                                st.session_state.storico_consegne[plesso]

                                            salva_storico_cloud(st.session_state.storico_consegne);
                                            salva_ritiri_cloud(st.session_state.storico_ritiri);
                                            st.rerun()

                                    if col_del.button("❌", key=f"del_h_{plesso}_{tipo}_{i}"):
                                        aggiungi_ritiri(plesso, tipo, [lib.copy()])
                                        per_tipo[tipo].pop(i)
                                        if not per_tipo[tipo]: del per_tipo[tipo]
                                        salva_storico_cloud(st.session_state.storico_consegne);
                                        salva_ritiri_cloud(st.session_state.storico_ritiri);
                                        st.rerun()

    if st.button("⬅️ Torna al Menu"): st.session_state.pagina = "Inserimento"; st.rerun()

# ==============================================================================
# BLOCCO 11: PAGINA NUOVO LIBRO (CATALOGO)
# ==============================================================================
elif st.session_state.pagina == "NuovoLibro":
    st.subheader("🆕 Aggiungi nuovo titolo")

    # Inizializza chiavi session_state per pulire i campi
    if "nuovo_libro_salvato" not in st.session_state:
        st.session_state.nuovo_libro_salvato = False

    with st.container(border=True):
        nt = st.text_input("Titolo Libro", key="input_titolo",
                           value="" if st.session_state.nuovo_libro_salvato else st.session_state.get("input_titolo",
                                                                                                      ""))
        col1, col2, col3 = st.columns(3)
        m_val = col1.text_input("Materia", key="input_materia",
                                value="" if st.session_state.nuovo_libro_salvato else st.session_state.get(
                                    "input_materia", ""))
        e_val = col2.text_input("Editore", key="input_editore",
                                value="" if st.session_state.nuovo_libro_salvato else st.session_state.get(
                                    "input_editore", ""))
        a_val = col3.text_input("Agenzia", key="input_agenzia",
                                value="" if st.session_state.nuovo_libro_salvato else st.session_state.get(
                                    "input_agenzia", ""))

        if st.button("✅ SALVA", use_container_width=True, type="primary"):
            if not nt or not m_val or not e_val:
                st.warning("⚠️ Inserisci almeno Titolo, Materia ed Editore!")
            else:
                if aggiungi_libro_a_excel(nt, m_val, e_val, a_val):
                    st.session_state.nuovo_libro_salvato = True
                    st.success(f"✅ Libro **'{nt}'** aggiunto con successo!")
                    st.rerun()
                else:
                    st.error("❌ Errore durante il salvataggio del libro.")

        # Reset flag dopo rerun
        if st.session_state.nuovo_libro_salvato:
            st.session_state.nuovo_libro_salvato = False
# ------------------------------------------------------------------------------

# ==============================================================================
# BLOCCO 12: PAGINA MODIFICA LIBRO (CATALOGO)
# ==============================================================================
elif st.session_state.pagina == "ModificaLibro":
    st.subheader("✏️ Modifica libro in Catalogo")

    sh = connetti_google_sheets()
    if not sh:
        st.error("❌ Impossibile connettersi a Google Sheets")
    else:
        foglio_catalogo = sh.worksheet("Catalogo")
        valori = foglio_catalogo.get_all_values()

        if not valori or len(valori) < 2:
            st.info("Catalogo vuoto.")
        else:
            righe = valori[1:]

            def _s(v):
                return str(v).strip()

            def _su(v):
                return _s(v).upper()

            records = []
            for i, r in enumerate(righe, start=2):
                titolo = _s(r[0]) if len(r) > 0 else ""
                materia = _s(r[1]) if len(r) > 1 else ""
                editore = _s(r[2]) if len(r) > 2 else ""
                agenzia = _s(r[3]) if len(r) > 3 else ""
                if titolo:
                    records.append({"row": i, "t": titolo, "m": materia, "e": editore, "a": agenzia})

            if not records:
                st.info("Catalogo vuoto.")
            else:
                titoli_unici = sorted(list({r["t"] for r in records}))
                titolo_sel = st.selectbox("Seleziona titolo da modificare:", ["- SELEZIONA -"] + titoli_unici)

                if titolo_sel != "- SELEZIONA -":
                    candidati = [r for r in records if _su(r["t"]) == _su(titolo_sel)]
                    opzioni = [f"{c['t']} | {c['m']} | {c['e']} | {c['a']} (riga {c['row']})" for c in candidati]
                    scelta = st.selectbox("Record:", opzioni) if len(opzioni) > 1 else opzioni[0]
                    idx = opzioni.index(scelta) if len(opzioni) > 1 else 0
                    rec = candidati[idx]

                    with st.container(border=True):
                        nt = st.text_input("Titolo Libro", value=rec["t"], key=f"mod_t_{rec['row']}")
                        col1, col2, col3 = st.columns(3)
                        m_val = col1.text_input("Materia", value=rec["m"], key=f"mod_m_{rec['row']}")
                        e_val = col2.text_input("Editore", value=rec["e"], key=f"mod_e_{rec['row']}")
                        a_val = col3.text_input("Agenzia", value=rec["a"], key=f"mod_a_{rec['row']}")

                        b1, b2 = st.columns(2)
                        if b1.button("✅ SALVA MODIFICHE", use_container_width=True, type="primary", key=f"btn_save_{rec['row']}"):
                            if not nt or not m_val or not e_val:
                                st.warning("⚠️ Inserisci almeno Titolo, Materia ed Editore!")
                            else:
                                if aggiorna_libro_catalogo(rec["row"], nt, m_val, e_val, a_val):
                                    st.success("✅ Libro aggiornato con successo!")
                                    st.rerun()
                                else:
                                    st.error("❌ Errore durante l'aggiornamento del libro.")

                        del_flag_key = f"del_conf_{rec['row']}"
                        if b2.button("🗑️ ELIMINA LIBRO", use_container_width=True, key=f"btn_del_{rec['row']}"):
                            st.session_state[del_flag_key] = True
                            st.rerun()

                        if st.session_state.get(del_flag_key):
                            st.warning("Confermi eliminazione del libro dal catalogo?")
                            c1, c2 = st.columns(2)
                            if c1.button("✅ CONFERMA ELIMINA", use_container_width=True, key=f"btn_del_ok_{rec['row']}"):
                                if elimina_libro_catalogo(rec["row"]):
                                    st.success("✅ Libro eliminato con successo!")
                                    st.session_state.pop(del_flag_key, None)
                                    st.rerun()
                            if c2.button("Annulla", use_container_width=True, key=f"btn_del_no_{rec['row']}"):
                                st.session_state.pop(del_flag_key, None)
                                st.rerun()

# ------------------------------------------------------------------------------

# ------------------------------------------------------------------------------
# BLOCCO 13: PAGINA REGISTRO E MOTORE DI RICERCA
# ==============================================================================
elif st.session_state.pagina == "Registro":
    st.subheader("📑 Registro Completo")
    if os.path.exists(DB_FILE):
        st.dataframe(pd.read_csv(DB_FILE), use_container_width=True)

elif st.session_state.pagina == "Ricerca":
    st.subheader("🔍 Motore di Ricerca Adozioni")
    if "r_attiva" not in st.session_state: st.session_state.r_attiva = False
    with st.container(border=True):
        r1c1, r1c2, r1c3 = st.columns(3)
        with r1c1:
            f_tit = st.multiselect("📕 Titolo", elenco_titoli, key="ft")
        with r1c2:
            f_age = st.multiselect("🤝 Agenzia", elenco_agenzie, key="fa")
        with r1c3:
            f_sag = st.selectbox("📚 Saggio", ["TUTTI", "SI", "NO"], key="fsag")
        r2c1, r2c2, r2c3 = st.columns(3)
        with r2c1:
            f_ple = st.multiselect("🏫 Plesso", ["NESSUNO"] + elenco_plessi, key="fp")
        with r2c2:
            f_mat = st.multiselect("📖 Materia", elenco_materie, key="fm")
        with r2c3:
            f_edi = st.multiselect("🏢 Editore", elenco_editori, key="fe")
        btn1, btn2, _ = st.columns([1, 1, 2])
        if btn1.button("🔍 AVVIA RICERCA", use_container_width=True, type="primary"): st.session_state.r_attiva = True
        if btn2.button("🧹 PULISCI", use_container_width=True, on_click=reset_ricerca): st.rerun()

    if st.session_state.r_attiva and os.path.exists(DB_FILE):
        df = pd.read_csv(DB_FILE).fillna("").astype(str)
        if f_ple: df = df[df["Plesso"].isin(f_ple)]
        if f_tit: df = df[df["Titolo"].isin(f_tit)]
        if f_age: df = df[df["Agenzia"].isin(f_age)]
        if f_mat: df = df[df["Materia"].isin(f_mat)]
        if f_edi: df = df[df["Editore"].isin(f_edi)]
        if f_sag != "TUTTI": df = df[df["Saggio Consegna"] == f_sag]
        if not df.empty:
            df_view = df.sort_index(ascending=False).copy()
            st.dataframe(df_view, use_container_width=True)
            somma = pd.to_numeric(df["N° sezioni"], errors='coerce').sum()
            
            tot_alunni = 0
            if "N° Alunni" in df.columns:
                tot_alunni = pd.to_numeric(df["N° Alunni"], errors='coerce').fillna(0).sum()
                
            st.markdown(f"""<div class="totale-box">
                🔢 Totale Classi: <b>{int(somma)}</b><br>
                👨‍🎓 Totale Alunni: <b>{int(tot_alunni)}</b>
                </div>""", unsafe_allow_html=True)
            
            out = io.BytesIO()
            try:
                df_export = df_view.copy()
                if "N° sezioni" in df_export.columns:
                    df_export["N° sezioni"] = pd.to_numeric(df_export["N° sezioni"], errors="coerce").fillna(0).astype(int)
                df_export.to_excel(out, index=False, sheet_name="Pivot Adozioni")
                st.download_button(
                    "📥 SCARICA EXCEL",
                    data=out.getvalue(),
                    file_name=f"pivot_adozioni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as ex:
                st.error(f"⚠️ Scarico Excel non disponibile: {ex}")
# ------------------------------------------------------------------------------
# =========================================================
# --- BLOCCO 12: PAGINA INSERIMENTO ADOZIONE ---
# INIZIO BLOCCO
# =========================================================
elif st.session_state.pagina == "Inserimento":
    st.subheader("Nuova Registrazione Adozione")
    if "form_id" not in st.session_state: st.session_state.form_id = 0
    
    if "prefill_adozione" not in st.session_state:
        st.session_state.prefill_adozione = {}

    if st.session_state.get("adozione_da_storico") and not st.session_state.get("prefill_adozione"):
        dati = st.session_state.adozione_da_storico
        st.markdown("### 🌟 Adozione da Collane Consegnate")
        st.write(f"**Plesso:** {dati.get('plesso','')} | **Titolo:** {dati.get('titolo','')} | **Editore:** {dati.get('editore','')}")
        with st.container(border=True):
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                n_sez_adott = st.number_input("Numero delle sezioni", min_value=1, value=1, key=f"imp_n_{st.session_state.form_id}")
            with c2:
                sez_lett_adott = st.text_input("Lettera sezione (es. A, B, C)", key=f"imp_sez_{st.session_state.form_id}")
            with c3:
                saggio_adott = st.selectbox("Saggio consegnato?", ["-", "NO", "SI"], key=f"imp_sag_{st.session_state.form_id}")
            with c4:
                n_alunni_adott = st.number_input("👨‍🎓 N° Alunni (opzionale)", min_value=0, value=0, key=f"imp_alunni_{st.session_state.form_id}")

            b1, b2 = st.columns(2)
            if b1.button("✅ CONFERMA DATI", type="primary", use_container_width=True, key=f"imp_ok_{st.session_state.form_id}"):
                st.session_state.prefill_adozione = {
                    "plesso": dati.get("plesso", ""),
                    "titolo": dati.get("titolo", ""),
                    "n_sez": n_sez_adott,
                    "sez_lett": sez_lett_adott,
                    "saggio": saggio_adott,
                    "n_alunni": n_alunni_adott if n_alunni_adott > 0 else 0,
                }
                st.session_state.adozione_da_storico = None
                st.session_state.form_id += 1
                st.rerun()
            if b2.button("⬅️ ANNULLA", use_container_width=True, key=f"imp_no_{st.session_state.form_id}"):
                st.session_state.adozione_da_storico = None
                st.session_state.pagina = "Storico"
                st.rerun()
        st.stop()

    prefill = st.session_state.get("prefill_adozione", {})
    
    with st.container(border=True):
        default_titolo_idx = 0
        if prefill.get("titolo") in elenco_titoli:
            default_titolo_idx = elenco_titoli.index(prefill["titolo"]) + 1

        titolo_scelto = st.selectbox("📕 SELEZIONA TITOLO", [""] + elenco_titoli, index=default_titolo_idx, key=f"tit_{st.session_state.form_id}")
        if titolo_scelto:
            info = catalogo[catalogo.iloc[:, 0] == titolo_scelto]
            if not info.empty:
                st.info(f"Materia: {info.iloc[0, 1]} | Editore: {info.iloc[0, 2]} | Agenzia: {info.iloc[0, 3]}")
        c1, c2, c3 = st.columns([2, 1, 1])
        with c1:
            default_plesso_idx = 0
            if prefill.get("plesso") in elenco_plessi:
                default_plesso_idx = elenco_plessi.index(prefill["plesso"]) + 1

            plesso = st.selectbox("🏫 Plesso", [""] + elenco_plessi, index=default_plesso_idx, key=f"ple_{st.session_state.form_id}")
            note = st.text_area("📝 Note", key=f"not_{st.session_state.form_id}", height=70)
        with c2:
            default_n_sez = prefill.get("n_sez", 1)
            n_sez = st.number_input("🔢 N° sezioni", min_value=1, value=default_n_sez, key=f"n_{st.session_state.form_id}")
            
            default_saggio_idx = 0
            if prefill.get("saggio") in ["-", "NO", "SI"]:
                default_saggio_idx = ["-", "NO", "SI"].index(prefill["saggio"])
                
            saggio = st.selectbox("📚 Saggio consegnato", ["-", "NO", "SI"], index=default_saggio_idx, key=f"sag_{st.session_state.form_id}")
        with c3:
            default_sez_lett = prefill.get("sez_lett", "")
            sez_lett = st.text_input("🔡 Lettera Sezione", value=default_sez_lett, key=f"sez_{st.session_state.form_id}")
            
            default_n_alunni = prefill.get("n_alunni", 0)
            n_alunni = st.number_input("👨‍🎓 N° Alunni (opzionale)", min_value=0, value=default_n_alunni, key=f"alunni_{st.session_state.form_id}")
            
        if st.button("💾 SALVA ADOZIONE", use_container_width=True, type="primary"):
            if titolo_scelto and plesso and saggio != "-":
                info = catalogo[catalogo.iloc[:, 0] == titolo_scelto]
                n_alunni_val = n_alunni if n_alunni > 0 else ""
                nuova_riga = pd.DataFrame([{
                    "Data": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "Plesso": plesso, "Materia": info.iloc[0, 1], "Titolo": titolo_scelto,
                    "Editore": info.iloc[0, 2], "Agenzia": info.iloc[0, 3], "N° sezioni": n_sez,
                    "Sezione": sez_lett.upper(), "Saggio Consegna": saggio, "N° Alunni": n_alunni_val, "Note": note
                }])
                df_attuale = pd.read_csv(DB_FILE) if os.path.exists(DB_FILE) else pd.DataFrame()
                df_finale = pd.concat([df_attuale, nuova_riga], ignore_index=True)
                df_finale.to_csv(DB_FILE, index=False)
                backup_su_google_sheets(df_finale)
                st.session_state.form_id += 1
                st.session_state.prefill_adozione = {}
                st.success("✅ Registrazione avvenuta con successo!")
                st.rerun()
            elif saggio == "-":
                st.error("⚠️ Devi specificare SI/NO!")
            else:
                st.error("⚠️ Seleziona Titolo e Plesso!")
# =========================================================
# FINE BLOCCO 12
# =========================================================
# ==============================================================================
# BLOCCO 11: PAGINA NUOVO LIBRO (CATALOGO)
# ==============================================================================
elif st.session_state.pagina == "NuovoLibro":
    st.subheader("🆕 Aggiungi nuovo titolo")
    with st.container(border=True):
        nt = st.text_input("Titolo Libro")
        col1, col2, col3 = st.columns(3)
        m_val = col1.text_input("Materia")
        e_val = col2.text_input("Editore")
        a_val = col3.text_input("Agenzia")
        if st.button("✅ SALVA", use_container_width=True, type="primary"):
            if nt and m_val and e_val:
                if aggiungi_libro_a_excel(nt, m_val, e_val, a_val):
                    st.success("Libro aggiunto!");
                    st.rerun()
# ------------------------------------------------------------------------------
# =========================================================
# --- BLOCCO 14: PAGINA MODIFICA (VERSIONE STABILE) ---
# =========================================================
elif st.session_state.pagina == "Modifica":
    st.subheader("✏️ Modifica o Cancella Adozioni")

    # Inizializziamo un contatore per il reset se non esiste
    if "reset_mod_ctr" not in st.session_state:
        st.session_state.reset_mod_ctr = 0

    if os.path.exists(DB_FILE):
        df_mod = pd.read_csv(DB_FILE).fillna("").astype(str)
        c_ric1, c_ric2 = st.columns(2)

        # Usiamo il contatore nella KEY per resettare tutto quando vogliamo
        with c_ric1:
            lista_plessi_db = sorted([x for x in df_mod["Plesso"].unique() if x != ""])
            p_cerca = st.selectbox("🔍 Filtra per Plesso", [""] + lista_plessi_db,
                                   key=f"p_mod_{st.session_state.reset_mod_ctr}")
        with c_ric2:
            lista_titoli_db = sorted([x for x in df_mod["Titolo"].unique() if x != ""])
            t_cerca = st.selectbox("🔍 Filtra per Titolo", [""] + lista_titoli_db,
                                   key=f"t_mod_{st.session_state.reset_mod_ctr}")

        if p_cerca or t_cerca:
            df_filtrato = df_mod.copy()
            if p_cerca: df_filtrato = df_filtrato[df_filtrato["Plesso"] == p_cerca]
            if t_cerca: df_filtrato = df_filtrato[df_filtrato["Titolo"] == t_cerca]

            if not df_filtrato.empty:
                for i in df_filtrato.index:
                    with st.container(border=True):
                        st.markdown(f"**Registrazione del {df_mod.at[i, 'Data']}**")
                        col1, col2, col3 = st.columns([2, 1, 1])
                        with col1:
                            try:
                                idx_p = elenco_plessi.index(df_mod.at[i, 'Plesso'])
                            except:
                                idx_p = 0
                            nuovo_plesso = st.selectbox(f"Plesso", elenco_plessi, index=idx_p, key=f"mp_{i}")
                            try:
                                idx_t = elenco_titoli.index(df_mod.at[i, 'Titolo'])
                            except:
                                idx_t = 0
                            nuovo_titolo = st.selectbox(f"Titolo Libro", elenco_titoli, index=idx_t, key=f"mt_{i}")
                            nuove_note = st.text_area("Note", value=df_mod.at[i, 'Note'], key=f"mnot_{i}", height=70)
                        with col2:
                            val_sez = int(float(df_mod.at[i, 'N° sezioni'])) if df_mod.at[i, 'N° sezioni'] else 1
                            nuovo_n_sez = st.number_input("N° sezioni", min_value=1, value=val_sez, key=f"mn_{i}")
                            nuova_sez_lett = st.text_input("Lettera Sezione", value=df_mod.at[i, 'Sezione'],
                                                           key=f"ms_{i}")
                        with col3:
                            attuale_sag = df_mod.at[i, 'Saggio Consegna']
                            idx_saggio = ["-", "NO", "SI"].index(attuale_sag) if attuale_sag in ["-", "NO", "SI"] else 0
                            nuovo_saggio = st.selectbox("Saggio consegnato", ["-", "NO", "SI"], index=idx_saggio,
                                                        key=f"msag_{i}")
                                                        
                            val_alunni = 0
                            if "N° Alunni" in df_mod.columns and df_mod.at[i, 'N° Alunni']:
                                try:
                                    val_alunni = int(float(df_mod.at[i, 'N° Alunni']))
                                except ValueError:
                                    pass
                            nuovo_n_alunni = st.number_input("👨‍🎓 N° Alunni", min_value=0, value=val_alunni, key=f"malunni_{i}")

                        b1, b2 = st.columns(2)
                        with b1:
                            if st.button("💾 AGGIORNA", key=f"upd_{i}", use_container_width=True, type="primary"):
                                if nuovo_saggio != "-":
                                    df_full = pd.read_csv(DB_FILE).fillna("").astype(str)
                                    info_new = catalogo[catalogo.iloc[:, 0] == nuovo_titolo]
                                    df_full.at[i, 'Plesso'] = nuovo_plesso
                                    df_full.at[i, 'Titolo'] = nuovo_titolo
                                    if not info_new.empty:
                                        df_full.at[i, 'Materia'] = info_new.iloc[0, 1]
                                        df_full.at[i, 'Editore'] = info_new.iloc[0, 2]
                                        df_full.at[i, 'Agenzia'] = info_new.iloc[0, 3]
                                    df_full.at[i, 'N° sezioni'] = nuovo_n_sez
                                    df_full.at[i, 'Sezione'] = nuova_sez_lett.upper()
                                    df_full.at[i, 'Saggio Consegna'] = nuovo_saggio
                                    df_full.at[i, 'Note'] = nuove_note
                                    df_full.at[i, 'N° Alunni'] = nuovo_n_alunni if nuovo_n_alunni > 0 else ""

                                    df_full.to_csv(DB_FILE, index=False)
                                    backup_su_google_sheets(df_full)

                                    # --- TRUCCO PER IL RESET: Incrementiamo il contatore ---
                                    st.session_state.reset_mod_ctr += 1

                                    st.success("Aggiornato!");
                                    st.rerun()
                        with b2:
                            if st.button("🗑️ ELIMINA", key=f"del_{i}", use_container_width=True):
                                df_full = pd.read_csv(DB_FILE).fillna("").astype(str)
                                df_full = df_full.drop(int(i))
                                df_full.to_csv(DB_FILE, index=False)
                                backup_su_google_sheets(df_full)
                                st.session_state.reset_mod_ctr += 1  # Reset anche qui
                                st.rerun()
# FINE BLOCCO 14
# =========================================================
# =========================================================
# --- BLOCCO 15: TABELLONE GENERALE (FORMA RETTANGOLARE) ---
# =========================================================
elif st.session_state.pagina == "Tabellone Stato":
    st.header("📊 Tabellone Avanzamento Plessi")

    if "storico_consegne" not in st.session_state:
        st.session_state.storico_consegne = carica_storico_cloud()
    if "storico_ritiri" not in st.session_state:
        st.session_state.storico_ritiri = carica_ritiri_cloud()

    # Mappatura Sigle Aggiornata
    mappa_sigle = {
        "LETTURE CLASSE PRIMA": "L1",
        "LETTURE CLASSE QUARTA": "L4",
        "SUSSIDIARI DISCIPLINE": "S4",
        "RELIGIONE": "R1\\4",
        "INGLESE CLASSE PRIMA": "E1",
        "INGLESE CLASSE QUARTA": "E4",
        "QUADERNI VACANZE CLASSE PRIMA": "V1",
        "QUADERNI VACANZE CLASSE SECONDA": "V2",
        "QUADERNI VACANZE CLASSE TERZA": "V3",
        "QUADERNI VACANZE CLASSE QUARTA": "V4",
        "QUADERNI VACANZE CLASSE QUINTA": "V5",
        "QUADERNI VACANZE INGLESE": "VE"
    }

    # Definizione ordine per righe rettangolo
    riga_sup_target = ["L1", "L4", "S4", "R1\\4", "E1", "E4"]
    riga_inf_target = ["V1", "V2", "V3", "V4", "V5", "VE"]

    elenco_totale = get_lista_plessi()
    consegnati = st.session_state.get("storico_consegne", {})
    ritirati = st.session_state.get("storico_ritiri", {})

    if not elenco_totale:
        st.warning("⚠️ Nessun plesso trovato.")
    else:
        # Statistiche
        n_tot = len(elenco_totale)
        n_ritirati_count = len([p for p in elenco_totale if p in ritirati and not consegnati.get(p)])
        n_consegnati_count = len([p for p in elenco_totale if p in consegnati])
        n_bianchi_count = n_tot - (len(set(consegnati.keys()) | set(ritirati.keys())))

        c1, c2, c3 = st.columns(3)
        c1.metric("⚪ DA INIZIARE", max(0, n_bianchi_count))
        c2.metric("🟠 DA RITIRARE", n_consegnati_count)
        c3.metric("🟢 COMPLETATI", n_ritirati_count)

        st.markdown("---")

        # Filtri
        f1, f2 = st.columns([2, 1])
        with f1:
            cerca_sel = st.selectbox("🔍 Cerca Plesso...", ["- TUTTI -"] + elenco_totale)
        with f2:
            filtro_stato = st.selectbox("📂 Filtra per Stato",
                                        ["TUTTI", "DA INIZIARE", "DA RITIRARE", "RITIRATI"])

        mostra = []
        for p in elenco_totale:
            if cerca_sel != "- TUTTI -" and p != cerca_sel: continue
            cat_attive = consegnati.get(p, {}).keys()
            ha_sigle = len(cat_attive) > 0
            e_ritirato = p in ritirati and not ha_sigle
            e_bianco = p not in consegnati and p not in ritirati

            if filtro_stato == "TUTTI":
                mostra.append(p)
            elif filtro_stato == "DA INIZIARE" and e_bianco:
                mostra.append(p)
            elif filtro_stato == "DA RITIRARE" and ha_sigle:
                mostra.append(p)
            elif filtro_stato == "RITIRATI" and e_ritirato:
                mostra.append(p)

        # Griglia Plessi
        if not mostra:
            st.info("ℹ️ Nessun plesso trovato.")
        else:
            n_col = 4
            for i in range(0, len(mostra), n_col):
                cols = st.columns(n_col)
                for j, plesso in enumerate(mostra[i:i + n_col]):

                    categorie_attive = consegnati.get(plesso, {}).keys()
                    sigle_attive = [mappa_sigle.get(cat, cat[:2]) for cat in categorie_attive]

                    bg, txt, lab, brd = ("#f8f9fa", "#333", "DA INIZIARE", "2px solid #dee2e6")

                    if plesso in ritirati and not sigle_attive:
                        bg, txt, lab, brd = ("#28a745", "#FFF", "✅ COMPLETATO", "2px solid #1e7e34")
                    elif sigle_attive:
                        bg, txt, lab, brd = ("#FF8C00", "#FFF", "🚚 IN CONSEGNA", "2px solid #e67e22")

                    # --- COSTRUZIONE RETTANGOLO SIGLE ---
                    html_blocco_sigle = ""
                    if sigle_attive:
                        def crea_span(s):
                            return f'''<span style="background: white; color: black; padding: 2px 6px; 
                                     border-radius: 4px; font-size: 13px; font-weight: 900; margin: 2px; 
                                     border: 2px solid #000; display: inline-block; min-width: 35px;
                                     box-shadow: 1px 1px 0px rgba(0,0,0,0.2);">{s}</span>'''


                        # Dividiamo le sigle presenti nelle due righe
                        riga1 = [crea_span(s) for s in riga_sup_target if s in sigle_attive]
                        riga2 = [crea_span(s) for s in riga_inf_target if s in sigle_attive]

                        html_r1 = f'<div style="margin-bottom: 2px;">{"".join(riga1)}</div>' if riga1 else ""
                        html_r2 = f'<div>{"".join(riga2)}</div>' if riga2 else ""

                        html_blocco_sigle = f'<div style="margin-top: 10px; display: flex; flex-direction: column; align-items: center;">{html_r1}{html_r2}</div>'

                    with cols[j]:
                        st.markdown(f"""
                            <div style="
                                background-color: {bg}; color: {txt}; border: {brd};
                                border-radius: 12px; padding: 15px 5px; margin-bottom: 20px;
                                text-align: center; min-height: 220px; display: flex;
                                flex-direction: column; justify-content: center; align-items: center;
                                box-shadow: 0px 4px 10px rgba(0,0,0,0.1);
                            ">
                                <div style="font-size: 18px; font-weight: 900; line-height: 1.1; text-transform: uppercase; margin-bottom: 5px;">
                                    {plesso}
                                </div>
                                <div style="font-size: 10px; font-weight: bold; text-transform: uppercase; letter-spacing: 1px; opacity: 0.9;">
                                    {lab}
                                </div>
                                {html_blocco_sigle}
                            </div>
                        """, unsafe_allow_html=True)

    st.markdown("---")
    if st.button("⬅️ Torna al Modulo Consegne", key="btn_back_tab_final"):
        st.session_state.pagina = "Consegne";
        st.rerun()
# =========================================================
# --- BLOCCO 16: RICERCA COLLANE E CONSEGNE ---
# =========================================================
elif st.session_state.pagina == "Ricerca Collane":
    st.subheader("🔍 Motore di Ricerca Collane Consegnate")

    # 1. Inizializziamo il contatore di reset se non esiste
    if "reset_collane" not in st.session_state:
        st.session_state.reset_collane = 0

    if "storico_consegne" not in st.session_state:
        st.session_state.storico_consegne = carica_storico_cloud()

    # 2. Trasformazione dati per la tabella
    righe_storico = []
    if st.session_state.storico_consegne:
        for plesso, categorie in st.session_state.storico_consegne.items():
            for cat, libri in categorie.items():
                for lib in libri:
                    righe_storico.append({
                        "Plesso": plesso,
                        "Tipologia": cat,
                        "Titolo": lib.get('t', ''),
                        "Editore": lib.get('e', ''),
                        "Quantità": lib.get('q', 0)
                    })

    df_collane = pd.DataFrame(righe_storico)

    if not df_collane.empty:
        # --- AREA FILTRI ---
        with st.container(border=True):
            c1, c2, c3 = st.columns(3)

            # Creiamo un suffisso dinamico basato sul contatore di reset
            suff = str(st.session_state.reset_collane)

            # Applichiamo il suffisso alle key (es: f_ple_0, f_ple_1...)
            f_ple = c1.multiselect("🏫 Filtra Plesso", sorted(df_collane["Plesso"].unique()), key="f_ple_" + suff)
            f_tip = c2.multiselect("📚 Filtra Tipologia", sorted(df_collane["Tipologia"].unique()), key="f_tip_" + suff)
            f_edi = c3.multiselect("🏢 Filtra Editore", sorted(df_collane["Editore"].unique()), key="f_edi_" + suff)

            # --- TASTO PULISCI CORRETTO ---
            if st.button("🧹 PULISCI TUTTI I FILTRI", use_container_width=True):
                # Invece di svuotare le liste, cambiamo il nome delle chiavi dei widget
                st.session_state.reset_collane += 1
                st.rerun()

        # 3. Applicazione filtri
        df_filtrato = df_collane.copy()
        if f_ple: df_filtrato = df_filtrato[df_filtrato["Plesso"].isin(f_ple)]
        if f_tip: df_filtrato = df_filtrato[df_filtrato["Tipologia"].isin(f_tip)]
        if f_edi: df_filtrato = df_filtrato[df_filtrato["Editore"].isin(f_edi)]

        # 4. Risultati e Totale
        totale_copie_collane = int(df_filtrato["Quantità"].sum())

        st.markdown(f"""
            <div style="padding:20px; background-color:#e8f0fe; border-radius:10px; border-left:8px solid #004a99; margin-bottom:20px;">
                <h3 style='margin:0; color:#004a99;'>Riepilogo Consegne</h3>
                <p style='font-size:24px; margin:5px 0 0 0;'>
                    Totale Libri Trovati: <b>{totale_copie_collane}</b>
                </p>
            </div>
        """, unsafe_allow_html=True)

        df_view = df_filtrato.copy()
        if not df_view.empty:
            df_view = df_view.sort_values(by=["Plesso", "Tipologia", "Titolo", "Editore"], kind="stable")

        st.dataframe(df_view, use_container_width=True, hide_index=True)
        out = io.BytesIO()
        try:
            df_export = df_view.copy()
            if "Quantità" in df_export.columns:
                df_export["Quantità"] = pd.to_numeric(df_export["Quantità"], errors="coerce").fillna(0).astype(int)
            df_export.to_excel(out, index=False, sheet_name="Ricerca Collane")
            st.download_button(
                "📥 SCARICA EXCEL",
                data=out.getvalue(),
                file_name=f"ricerca_collane_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as ex:
            st.error(f"⚠️ Scarico Excel non disponibile: {ex}")

    else:
        st.warning("⚠️ Non ci sono ancora dati nello storico delle consegne.")

# =========================================================
# --- BLOCCO 17: COLLANE RITIRATE ---
# =========================================================
elif st.session_state.pagina == "Ritirate":
    st.subheader("📦 Collane Ritirate")
    if "storico_ritiri" not in st.session_state:
        st.session_state.storico_ritiri = carica_ritiri_cloud()
    if not st.session_state.storico_ritiri:
        st.info("ℹ️ Nessuna collana risulta ritirata al momento.")
    else:
        elenco_plessi_ritiri = sorted(list(st.session_state.storico_ritiri.keys()))
        plessi_show = st.multiselect("🔍 Seleziona Plesso/i:", elenco_plessi_ritiri, key="sel_plessi_ritirati")
        if not plessi_show:
            st.info("Seleziona uno o più plessi per visualizzare le collane ritirate.")
        else:
            for plesso in plessi_show:
                with st.expander(f"🏫 PLESSO: {plesso.upper()}", expanded=True):
                    per_tipo = st.session_state.storico_ritiri.get(plesso, {})
                    tot_plesso = 0
                    open_tipo_key = f"open_tipo_ritirate_{plesso}"
                    if open_tipo_key not in st.session_state:
                        st.session_state[open_tipo_key] = None

                    st.markdown("""
                        <style>
                        div[data-testid="stButton"] > button {
                            display: flex !important;
                            justify-content: flex-start !important;
                            text-align: left !important;
                            padding-left: 15px !important;
                        }
                        div[data-testid="stButton"] > button > div {
                            width: 100% !important;
                            display: flex !important;
                            justify-content: flex-start !important;
                        }
                        div[data-testid="stButton"] > button > div > p {
                            width: 100% !important;
                            text-align: left !important;
                            margin: 0 !important;
                        }
                        </style>
                        """, unsafe_allow_html=True)

                    for tipo in ordina_tipologie(per_tipo.keys()):
                        is_open = st.session_state.get(open_tipo_key) == tipo
                        freccia = "🔽" if is_open else "▶️"

                        if st.button(f"{freccia} 📚 {tipo.upper()}", key=f"open_tipo_rit_{plesso}_{tipo}", use_container_width=True):
                            st.session_state[open_tipo_key] = None if is_open else tipo
                            st.rerun()

                        if is_open:
                            with st.container(border=True):
                                libri = per_tipo[tipo]
                                agg = {}
                                for lib in libri:
                                    key = (lib.get('t',''), lib.get('e',''))
                                    agg[key] = agg.get(key, 0) + int(lib.get('q', 0))
                                df_tip = pd.DataFrame([{"Titolo": k[0], "Editore": k[1], "Quantità": q} for k, q in agg.items()])
                                if not df_tip.empty:
                                    df_tip = df_tip.sort_values(by=["Titolo", "Editore"])
                                    st.dataframe(df_tip, use_container_width=True, hide_index=True)
                                    tot_tipo = int(df_tip["Quantità"].sum())
                                    tot_plesso += tot_tipo
                                    st.markdown(f"<div class='totale-box'>Totale tipologia: <b>{tot_tipo}</b></div>", unsafe_allow_html=True)
                    st.markdown(f"<div class='totale-box'>Totale ritiri plesso: <b>{tot_plesso}</b></div>", unsafe_allow_html=True)
